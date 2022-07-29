from openpyxl import load_workbook
import openpyxl
import random
from datetime import datetime
from dateutil.relativedelta import relativedelta


# def key_creator():
#     letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
#     numeral = '0123456789'
#
#     all_simbols = letters + numeral
#     length = 9
#     password = ''.join(random.sample(all_simbols, length))
#     password = password[:-5] + '-' + password[4:]
#     return password
#
#
# def date_and_time():
#     dt_now = datetime.datetime.now()
#     date = str(dt_now)[8:-16] + '.' + str(dt_now)[5:-19] + '.' + str(dt_now)[0:-22]
#     time = str(dt_now)[11:16]
#     answer = f'{date} - {time}'
#     return answer
#
# file = 'Ключи.xlsx'
# # wb = load_workbook(file)
# # ws = wb['data']
# #
# # # count_of_keys = int(input('Введите количество ключей: '))
# # # for i in range(count_of_keys):
# # #     ws.append([f'Ключ №{i+1}', f'{key_creator()}', f'{date_and_time()}'])
# # wb.close()
#
# reading = openpyxl.open(file, read_only = True)
# sheet = reading.active
# print(sheet.max_row)

# file = 'Ключи.xlsx'
# wookbook = openpyxl.load_workbook(file)
# search_text = input(str('Какой текст ищем: '))
# print('Ищем:', search_text)
# sheet = wookbook.active
# num = 2
# print(sheet[f'D{num}'].value)
# #
# # for i in range(0, worksheet.max_row):
# #     for col in worksheet.iter_cols(1, worksheet.max_column):
# #         # print(col[i].value, end="\t\t")
# #         if search_text in col[i].value:
# #             print(f'{i} ключ')
# #             print(col)
#
# rows = sheet.max_row
# cols = sheet.max_column
#
# for i in range(1, rows + 1):
#     string = ''
#     for j in range(1, cols + 1):
#         cell = sheet.cell(row = i, column = j)
#         string = string + str(cell.value) + ' '
#     if search_text in string:
#         print(string[38])
# date_after_month = datetime.today()+ relativedelta(months=1)
# print('Today: ',datetime.today().strftime('%d.%m.%Y'))
# print('After Month:', date_after_month.strftime('%d.%m.%Y'))

# file = 'data.xlsx'
# woorkbook = openpyxl.load_workbook(file)
# sheet = woorkbook.active
# rows = sheet.max_row
# cols = sheet.max_column
# disabled_c = 0
# active_c = 0
# deactivated_c = 0
#
# for i in range(1, rows + 1):
#     for j in range(1, cols + 1):
#         if str(sheet[f'D{i}'].value) == '0':
#             disabled_c += 1
#         if str(sheet[f'D{i}'].value) == '1':
#             active_c += 1
#         if str(sheet[f'D{i}'].value) == '2':
#             deactivated_c += 1

# def date_and_time():
#     dt_now = datetime.now()
#     date = str(dt_now)[8:-16] + '.' + str(dt_now)[5:-19] + '.' + str(dt_now)[0:-22]
#     time = str(dt_now)[11:16]
#     answer = f'{date} - {time}'
#     return answer
#
#
# print(date_and_time())