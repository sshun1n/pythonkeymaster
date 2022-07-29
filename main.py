from PyQt5.QtWidgets import QApplication, QMainWindow
from dateutil.relativedelta import relativedelta
from PyQt5 import QtWidgets, QtCore, QtGui
from openpyxl import load_workbook
from design import Ui_MainWindow
from datetime import datetime
from PyQt5.QtGui import QIcon
import openpyxl
import random
import sys
import os


class Generator(QtWidgets.QMainWindow):
    def __init__(self):
        super(Generator, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        file_dir = 'C:\Program Files (x86)\Key Master by sshunin'
        if 'Key Master by sshunin' not in os.listdir('C:\Program Files (x86)'):
            os.mkdir(file_dir)
            wb = openpyxl.Workbook()
            sheet = wb['Sheet']
            sheet['A1'] = 'Номер ключа'
            sheet['B1'] = 'Ключ'
            sheet['C1'] = 'Дата генерации'
            sheet['D1'] = 'Статус'
            sheet['E1'] = 'Срок действия'
            sheet['F1'] = 'Агент'
            sheet['G1'] = 'Дата деактивации'
            sheet['H1'] = 'Сумма заказа агента'
            sheet['I1'] = 'Клиент'
            sheet['J1'] = 'Сумма заказа клиента'
            wb.save('C:\Program Files (x86)\Key Master by sshunin\data.xlsx')
            cfg_file = open("C:\Program Files (x86)\Key Master by sshunin\cfg.txt", "w+")
            cfg_file.write(f"cfg\nStartWindowId: 0\nSaveAllErrors: 0")
            cfg_file.close()
            os.mkdir('C:\Program Files (x86)\Key Master by sshunin\log')
            self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] Created main dir\n', True)

        elif 'data.xlsx' not in os.listdir(file_dir):
            wb = openpyxl.Workbook()
            sheet = wb['Sheet']
            sheet['A1'] = 'Номер ключа'
            sheet['B1'] = 'Ключ'
            sheet['C1'] = 'Дата генерации'
            sheet['D1'] = 'Статус'
            sheet['E1'] = 'Срок действия'
            sheet['F1'] = 'Агент'
            sheet['G1'] = 'Дата деактивации'
            sheet['H1'] = 'Сумма заказа агента'
            sheet['I1'] = 'Клиент'
            sheet['J1'] = 'Сумма заказа клиента'
            wb.save('C:\Program Files (x86)\Key Master by sshunin\data.xlsx')
            cfg_file = open("C:\Program Files (x86)\Key Master by sshunin\cfg.txt", "w+")
            cfg_file.write(f"cfg\nStartWindowId: 0\nSaveAllErrors: 0")
            cfg_file.close()

        if 'cfg.txt' not in os.listdir(file_dir):
            self.cfg_writer()

        else:
            if 'log' not in os.listdir(file_dir):
                os.mkdir('C:\Program Files (x86)\Key Master by sshunin\log')
                self.get_cfg()

            else:
                self.get_cfg()
            #     # log_file.write(f"cfg\nStartWindowId: 0\nSaveAllErrors: 0")
            #     log_file.close()

    def get_cfg(self):
        self.all_hide()

        cfg = open("C:\Program Files (x86)\Key Master by sshunin\cfg.txt", "r")
        settings_list = cfg.readlines()
        start_window_id = settings_list[1]
        save_all_errors = settings_list[2]
        start_window_id = int(start_window_id[-2])
        save_all_errors = int(save_all_errors[-1])

        self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] Started\n', True)

        if start_window_id == 0:
            self.init_UI()
            self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] window{start_window_id} opened successfully\n', False)

        elif start_window_id == 1:
            self.key_checker_redirect()
            self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] window{start_window_id} opened successfully\n', False)

        elif start_window_id == 2:
            self.generator_redirect()
            self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] window{start_window_id} opened successfully\n', False)

        else:
            self.all_hide()
            self.ui.eror_table.setText('Ошибка:  Файл настроек cfg.txt несиправен\n(StartWindowId: unknown argument)\n\nНеобходимо сбросить настройки')
            self.ui.eror_table.show()
            self.ui.eror_table_create_button.setText('Сброс')
            self.ui.eror_table_create_button.show()
            self.ui.eror_table_create_button.clicked.connect(lambda: [self.cfg_writer(), self.init_UI(),
                                                                      self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] Error: cfg incorrect (window{start_window_id})\n', False),
                                                                      self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] cfg was rewrote successfully\n', False)])

    def log_saver(self, func, new_file):
        cfg = open("C:\Program Files (x86)\Key Master by sshunin\cfg.txt", "r")
        settings_list = cfg.readlines()
        save_all_errors = settings_list[2]
        save_err = int(save_all_errors[-1])

        if save_err == 1:
            pass

        elif save_err == 0:
            if new_file:
                log_file = open(f"C:\Program Files (x86)\Key Master by sshunin\log\log [{datetime.today().strftime('%H.%M.%d.%m')}].txt", 'w+')
                log_file.write(func)
                log_file.close()
            if new_file is False:
                file = os.listdir('C:\Program Files (x86)\Key Master by sshunin\log')
                log_file = open(f"C:\Program Files (x86)\Key Master by sshunin\log\{file[-1]}", 'a')
                last_l = open(f"C:\Program Files (x86)\Key Master by sshunin\log\{file[-1]}", 'r')
                last_line = last_l.readlines()
                last_line = last_line[-1]
                if func[10:] != last_line[10:]:
                    log_file.write(func)
                    log_file.close()
                else:
                    pass

    def all_hide(self):
        self.ui.agen_label.hide()
        self.ui.date_label.hide()
        self.ui.key_val_label.hide()
        self.ui.key_status_label.hide()
        self.ui.status_button.hide()
        self.ui.save_button.hide()
        self.ui.agent_name_inp.hide()
        self.ui.agent_name_label.hide()
        self.ui.agent_name_exit_button.hide()
        self.ui.agent_name_inp_value.hide()
        self.ui.agent_name_vaule_label.hide()
        self.ui.deadline_label.hide()
        self.ui.deadline_slider.hide()
        self.ui.deadline_inp_label.hide()
        self.ui.deadline_label_2.hide()
        self.ui.deactivate_label.hide()
        self.ui.deactivate_button.hide()
        self.ui.key_status_label_2.hide()
        self.ui.deactivate_exit_button.hide()
        self.ui.key_generator_all_label.hide()
        self.ui.key_generator_label.hide()
        self.ui.key_generator_amount_inp.hide()
        self.ui.key_generator_generate_button.hide()
        self.ui.key_generator_disabled_label.hide()
        self.ui.key_generator_deactivated_label.hide()
        self.ui.key_generator_active_label.hide()
        self.ui.key_generator_textBrowser.hide()
        self.ui.eror_table.hide()
        self.ui.eror_table_create_button.hide()
        self.ui.get_random_key_label.hide()
        self.ui.get_random_key_button.hide()
        self.ui.key_inp.hide()
        self.ui.key_inp_button.hide()
        self.ui.key_inp_label.hide()
        self.ui.generator_button.hide()
        self.ui.checker_button.hide()
        self.ui.settings_label.hide()
        self.ui.settings_button.hide()
        self.ui.settings_main_window_label.hide()
        self.ui.settings_main_window_button.hide()
        self.ui.settings_main_window_value.hide()
        self.ui.settings_default_button.hide()
        self.ui.settings_default_label.hide()
        self.ui.settings_delete_table_button.hide()
        self.ui.settings_delete_table_label.hide()
        self.ui.settings_save_all_errors_label.hide()
        self.ui.settings_save_all_errors_button.hide()
        self.ui.settings_open_table_button.hide()
        self.ui.agen_label_2.hide()
        self.ui.agen_label_3.hide()
        self.ui.date_label_2.hide()
        self.ui.client_name_out.hide()
        self.ui.client_name_val_out.hide()
        self.ui.deadline_label_3.hide()

    def create_dir(self):
        self.all_hide()

        file = 'C:\Program Files (x86)\Key Master by sshunin\data.xlsx'
        file_dir = 'C:\Program Files(x86)\Key Master by sshunin '

        def create_xlsx():
            wb = openpyxl.Workbook()
            sheet = wb['Sheet']
            sheet['A1'] = 'Номер ключа'
            sheet['B1'] = 'Ключ'
            sheet['C1'] = 'Дата генерации'
            sheet['D1'] = 'Статус'
            sheet['E1'] = 'Срок действия'
            sheet['F1'] = 'Агент'
            sheet['G1'] = 'Дата деактивации'
            sheet['H1'] = 'Сумма заказа агента'
            sheet['I1'] = 'Клиент'
            sheet['J1'] = 'Сумма заказа клиента'
            wb.save(file)
            cfg_file = open("C:\Program Files (x86)\Key Master by sshunin\cfg.txt", "w+")
            cfg_file.write(f"cfg\nStartWindowId: 0\nSaveAllErrors: 0")
            cfg_file.close()

        self.ui.eror_table.setText('Ошибка:  Таблица с ключами не обнаружена. Создать таблицу?')
        self.ui.eror_table.show()
        self.ui.eror_table_create_button.setText('Создать таблицу')
        self.ui.eror_table_create_button.show()
        self.ui.eror_table_create_button.clicked.connect(lambda: [create_xlsx(), self.init_redirect(), self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] Table was created successfully\n', False)])

    def change_main_window(self):
        if self.ui.settings_main_window_button.value() == 1:
            self.ui.settings_main_window_value.setText('Главное меню')

        if self.ui.settings_main_window_button.value() == 2:
            self.ui.settings_main_window_value.setText('Проверка ключей')

        if self.ui.settings_main_window_button.value() == 3:
            self.ui.settings_main_window_value.setText('Генератор')

    def settings(self):
        self.all_hide()
        cfg = open("C:\Program Files (x86)\Key Master by sshunin\cfg.txt", "r")
        settings_list = cfg.readlines()
        start_window_id = settings_list[1]
        save_all_errors = settings_list[2]
        start_window_id = int(start_window_id[-2])
        save_all_errors = int(save_all_errors[-1])

        self.ui.settings_label.show()
        self.ui.settings_main_window_label.show()
        self.ui.settings_open_table_button.show()
        self.ui.settings_open_table_button.clicked.connect(lambda: os.startfile(r'C:\Program Files (x86)\Key Master by sshunin\data.xlsx'))
        self.ui.settings_main_window_button.show()
        self.ui.settings_main_window_button.valueChanged.connect(self.change_main_window)
        if start_window_id == 0:
            self.ui.settings_main_window_value.setText('Главное меню')
            self.ui.settings_main_window_button.setValue(0)

        elif start_window_id == 1:
            self.ui.settings_main_window_value.setText('Проверка ключей')
            self.ui.settings_main_window_button.setValue(1)

        elif start_window_id == 2:
            self.ui.settings_main_window_value.setText('Генератор')
            self.ui.settings_main_window_button.setValue(2)

        self.ui.settings_main_window_value.show()
        self.ui.settings_default_button.show()
        self.ui.settings_default_label.show()
        self.ui.settings_default_button.clicked.connect(self.default_warning)
        self.ui.settings_delete_table_button.show()
        self.ui.settings_delete_table_button.clicked.connect(self.delete_warning)
        self.ui.settings_delete_table_label.show()
        self.ui.settings_save_all_errors_label.show()
        if save_all_errors == 0:
            save_errors_val = self.ui.settings_save_all_errors_button.setChecked(True)

        else:
            save_errors_val = self.ui.settings_save_all_errors_button.setChecked(False)
        self.ui.settings_save_all_errors_button.show()
        self.ui.agent_name_exit_button.show()
        self.ui.agent_name_exit_button.clicked.connect(lambda: self.init_UI())
        self.ui.save_button.show()
        self.ui.save_button.clicked.connect(self.cfg_settings_changer)

    def default_warning(self):
        self.all_hide()
        self.ui.deactivate_label.setText(f'Вы действительно хотите сбросить настройки?')
        self.ui.deactivate_label.show()
        self.ui.deactivate_exit_button.show()
        self.ui.deactivate_button.setText(f'Сброс')
        self.ui.deactivate_button.show()
        self.ui.deactivate_button.clicked.connect(lambda: [self.cfg_writer(), self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] Settings was vanished\n', False)])

    def delete_warning(self):
        self.all_hide()
        self.ui.deactivate_label.setText(f'Вы действительно хотите удалить таблицу?')
        self.ui.deactivate_label.show()
        self.ui.deactivate_exit_button.show()
        self.ui.deactivate_button.setText(f'Удалить')
        self.ui.deactivate_button.show()
        self.ui.deactivate_button.clicked.connect(self.init_UI)
        self.ui.deactivate_button.clicked.connect(lambda: [os.remove('C:\Program Files (x86)\Key Master by sshunin\data.xlsx'), self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] Table was deleted\n', False)])

    def random_key_finder(self):
        file = 'C:\Program Files (x86)\Key Master by sshunin\data.xlsx'
        woorkbook = openpyxl.load_workbook(file)
        free_keys = []
        sheet = woorkbook.active
        rows = sheet.max_row
        for i in range(1, rows + 1):
            if str(sheet[f'D{i}'].value) == '0':
                free_keys.append(sheet[f'B{i}'].value)

        self.ui.get_random_key_label.setText(random.choice(free_keys))

    def init_redirect(self):
        self.ui.eror_table.setText('Таблица успешно создана!')
        self.ui.eror_table.show()
        self.ui.eror_table_create_button.setText('Меню')
        self.ui.eror_table_create_button.show()
        self.ui.eror_table_create_button.clicked.connect(self.init_UI)

    def cfg_writer(self):
        cfg_file = open("C:\Program Files (x86)\Key Master by sshunin\cfg.txt", "w+")
        cfg_file.write(f"cfg\nStartWindowId: 0\nSaveAllErrors: 0")
        cfg_file.close()
        self.init_UI()

    def cfg_settings_changer(self):
        main_window = int(self.ui.settings_main_window_button.value())
        save_errors_val = 0

        if self.ui.settings_save_all_errors_button.isChecked() is True:
            save_errors_val = 0

        elif self.ui.settings_save_all_errors_button.isChecked() is False:
            save_errors_val = 1

        cfg_file = open("C:\Program Files (x86)\Key Master by sshunin\cfg.txt", "w+")
        cfg_file.write(f"cfg\nStartWindowId: {main_window - 1}\nSaveAllErrors: {save_errors_val}")
        cfg_file.close()
        self.init_UI()

    def init_UI(self):
        self.all_hide()
        self.ui.exit_button.hide()
        self.ui.settings_button.show()
        self.ui.settings_button.clicked.connect(lambda: self.settings())

        file = 'C:\Program Files (x86)\Key Master by sshunin\data.xlsx'
        file_dir = 'C:\Program Files (x86)\Key Master by sshunin'
        wookbook = openpyxl.load_workbook(file)
        sheet = wookbook.active
        rows = sheet.max_row
        disabled_c = 0
        for i in range(1, rows + 1):
            if str(sheet[f'D{i}'].value) == '0':
                disabled_c += 1

        if 'Key Master by sshunin' not in os.listdir('C:\Program Files (x86)'):
            os.mkdir(file_dir)
            self.create_dir()

        elif 'data.xlsx' not in os.listdir(file_dir):
            self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] Error: No table or main_dir\n', True)
            self.create_dir()

        else:
            self.ui.generator_button.show()
            self.ui.generator_button.clicked.connect(lambda: self.generator_redirect())
            self.ui.checker_button.show()
            self.ui.checker_button.clicked.connect(lambda: self.key_checker_redirect())
            if rows != 1:
                if disabled_c != 0:
                    self.ui.get_random_key_label.show()
                    self.ui.get_random_key_button.show()
                    self.ui.get_random_key_button.clicked.connect(self.random_key_finder)

    def generator_redirect(self):
        self.all_hide()
        self.ui.exit_button.show()
        self.ui.exit_button.clicked.connect(lambda: self.init_UI())
        self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] Redirected to generator\n', False)

        file = 'C:\Program Files (x86)\Key Master by sshunin\data.xlsx'
        woorkbook = openpyxl.load_workbook(file)
        sheet = woorkbook.active
        rows = sheet.max_row
        cols = sheet.max_column
        disabled_c = 0
        active_c = 0
        deactivated_c = 0

        for i in range(1, rows + 1):
            if str(sheet[f'D{i}'].value) == '0':
                disabled_c += 1
            if str(sheet[f'D{i}'].value) == '1':
                active_c += 1
            if str(sheet[f'D{i}'].value) == '2':
                deactivated_c += 1

        self.ui.key_generator_label.show()
        self.ui.key_generator_all_label.setText(f'Сгенерированно: {rows - 1}')
        self.ui.key_generator_all_label.show()
        self.ui.key_generator_active_label.setText(f'Активно: {active_c}')
        self.ui.key_generator_active_label.show()
        self.ui.key_generator_disabled_label.setText(f'Неактивно: {disabled_c}')
        self.ui.key_generator_disabled_label.show()
        self.ui.key_generator_deactivated_label.setText(f'Использованно: {deactivated_c}')
        self.ui.key_generator_deactivated_label.show()
        self.ui.key_generator_textBrowser.show()
        self.ui.key_generator_amount_inp.setText('0')
        self.ui.key_generator_amount_inp.show()
        self.ui.key_generator_generate_button.show()
        self.ui.key_generator_generate_button.clicked.connect(lambda: self.generator())

    def generator(self):
        self.all_hide()
        input_amount = int(self.ui.key_generator_amount_inp.text())
        if input_amount == 0:
            self.ui.deactivate_label.show()
            self.ui.deactivate_label.setText(f'Количество ключей должно быть больше нуля')
            self.ui.agent_name_exit_button.setText('Отмена')
            self.ui.agent_name_exit_button.show()
            self.ui.agent_name_exit_button.clicked.connect(self.init_UI)
            self.ui.save_button.setText('Ок')
            self.ui.save_button.show()
            self.ui.save_button.clicked.connect(self.generator_redirect)

        else:
            def key_creator():
                letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
                numeral = '0123456789'

                all_simbols = letters + numeral
                length = 9
                password = ''.join(random.sample(all_simbols, length))
                password = password[:-5] + '-' + password[4:]
                return password

            def date_and_time():
                dt_now = datetime.now()
                date = str(dt_now)[8:-16] + '.' + str(dt_now)[5:-19] + '.' + str(dt_now)[0:-22]
                time = str(dt_now)[11:16]
                answer = f'{date} - {time}'
                return answer

            file = 'C:\Program Files (x86)\Key Master by sshunin\data.xlsx'
            wb = load_workbook(file)
            ws = wb['Sheet']
            reading = openpyxl.open(file, read_only=True)
            sheet = reading.active
            rows_amount = sheet.max_row
            for i in range(input_amount):
                key = key_creator()
                self.log_saver(
                    f'[{datetime.today().strftime("%H:%M:%S")}] Generated key: {key}\n', False)
                ws.append([f'Ключ №{rows_amount}', f'{key}', f'{date_and_time()}', 0])
                rows_amount += 1
            wb.save(file)
            wb.close()

            self.ui.deactivate_label.show()
            self.ui.deactivate_label.setText(f'Удачно сгенерированно {input_amount} ключей')
            self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] Successfully generated {input_amount} keys\n', False)

    def key_checker_redirect(self):
        def auto_capital():
            text = self.ui.key_inp.text()
            self.ui.key_inp.setText(text.upper())

        self.all_hide()
        self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] Redirected to key checker\n', False)
        self.ui.exit_button.show()
        self.ui.exit_button.clicked.connect(lambda: [self.init_UI(), self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] Redirected to menu (exit button pressed)\n', False)])
        self.ui.key_inp.show()
        self.ui.key_inp_button.show()
        self.ui.key_inp_label.show()
        self.ui.key_inp.textChanged.connect(lambda: auto_capital())
        self.ui.key_inp_button.clicked.connect(lambda: [self.key_checker(), self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] Checking key {self.ui.key_inp.text()}\n', False)])

        file = 'C:\Program Files (x86)\Key Master by sshunin\data.xlsx'
        wookbook = openpyxl.load_workbook(file)
        sheet = wookbook.active
        rows = sheet.max_row
        if rows == 1:
            self.all_hide()
            self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] Error: No keys in table\n', False)
            self.ui.eror_table.setText('Ошибка:  В таблице отсутствуют ключи. Необходимо сгенерировать хотя бы один ключ!')
            self.ui.eror_table.show()
            self.ui.eror_table_create_button.setText('Генератор')
            self.ui.eror_table_create_button.show()
            self.ui.eror_table_create_button.clicked.connect(lambda: self.generator_redirect())

    def key_checker(self):
        def name_of_agent(key, rows):
            self.all_hide()

            self.ui.key_inp_label.show()
            self.ui.key_inp.show()
            self.ui.key_inp_button.show()
            self.ui.key_val_label.show()
            self.ui.key_val_label.setText(f'Ключ: {key_input}')
            self.ui.key_status_label.setText(f'{key}')
            self.ui.key_status_label.show()
            self.ui.status_button.setText(key_status_button)
            self.ui.status_button.show()
            self.ui.agent_name_inp.show()
            self.ui.agent_name_label.show()
            self.ui.agent_name_vaule_label.show()
            self.ui.agent_name_inp_value.show()
            self.ui.agent_name_exit_button.show()
            self.ui.deadline_slider.show()
            self.ui.deadline_inp_label.show()
            self.ui.deadline_label_2.show()
            self.ui.save_button.show()
            self.ui.deadline_slider.valueChanged.connect(self.update_label)

            self.ui.save_button.clicked.connect(lambda: [self.update_xls(rows), self.key_checker()])

        def get_client_name(rows, key):
            self.all_hide()

            self.ui.key_inp_label.show()
            self.ui.key_inp.show()
            self.ui.key_inp_button.show()
            self.ui.key_val_label.show()
            self.ui.key_val_label.setText(f'Ключ: {key_input}')
            self.ui.key_status_label.setText(f'{key}')
            self.ui.key_status_label.show()
            self.ui.status_button.setText(key_status_button)
            self.ui.status_button.show()
            self.ui.agent_name_inp.setText('')
            self.ui.agent_name_inp.show()
            self.ui.agent_name_label.setText('Введите имя клиента:')
            self.ui.agent_name_label.show()
            self.ui.agent_name_vaule_label.setText('Введите сумму заказа:')
            self.ui.agent_name_vaule_label.show()
            self.ui.agent_name_inp_value.setText('')
            self.ui.agent_name_inp_value.show()
            self.ui.agent_name_exit_button.show()
            self.ui.save_button.show()

            self.ui.save_button.clicked.connect(lambda: deactivate(rows))

        key_input = self.ui.key_inp.text()


        file = 'C:\Program Files (x86)\Key Master by sshunin\data.xlsx'
        wookbook = openpyxl.load_workbook(file)
        sheet = wookbook.active
        rows = sheet.max_row
        cols = sheet.max_column

        for i in range(1, rows + 1):
            string = ''
            for j in range(1, cols + 1):
                cell = sheet.cell(row=i, column=j)
                string = string + str(cell.value) + ' '
            if key_input in string:
                rw = i
                self.ui.key_val_label.setText(f'Ключ: {key_input}')
                self.ui.key_val_label.show()
                status = str(sheet[f'D{i}'].value)

                def deactivate(row):
                    self.all_hide()

                    self.ui.deactivate_label.setText(f'Вы действительно хотите деактивировать ключ {key_input}?')
                    self.ui.deactivate_label.show()
                    self.ui.deactivate_exit_button.show()
                    self.ui.deactivate_button.show()

                    self.ui.deactivate_exit_button.clicked.connect(self.key_checker)
                    self.ui.deactivate_button.clicked.connect(lambda: [change_status(row, 2), wookbook.save(file), self.ui.key_inp.setText(''), self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] Key {key_input} was successfully deactivated\n', False)])
                    self.ui.deactivate_button.clicked.connect(self.all_hide)
                    self.ui.deactivate_button.clicked.connect(lambda: self.ui.deactivate_label.show())
                    self.ui.deactivate_button.clicked.connect(self.init_UI)

                def change_status(row, val):
                    if val == 2:
                        sheet[f'D{row}'] = 2
                        sheet[f'G{row}'] = str(datetime.today().strftime("%d.%m.%Y"))
                        sheet[f'I{row}'] = str(self.ui.agent_name_inp.text())
                        sheet[f'J{row}'] = int(self.ui.agent_name_inp_value.text())
                    if val == 1:
                        sheet[f'D{row}'].value = 1

                if status == '2':
                    self.all_hide()

                    self.ui.key_inp_label.show()
                    self.ui.key_inp.show()
                    self.ui.key_inp_button.show()
                    self.ui.key_val_label.show()
                    self.ui.key_status_label_2.setText(f'Деактивирован {sheet[f"G{rw}"].value}')
                    self.ui.key_status_label_2.show()
                    self.ui.agen_label_2.setText(f' Агент: {sheet[f"F{rw}"].value}')
                    self.ui.agen_label_2.show()
                    self.ui.date_label_2.setText(f" Дата генерации: {sheet[f'C{rw}'].value}")
                    self.ui.date_label_2.show()
                    self.ui.deadline_label_3.setText(f' Срок действия: {sheet[f"E{rw}"].value}')
                    self.ui.deadline_label_3.show()
                    self.ui.agen_label_3.setText(f" Сумма заказа агента: {sheet[f'H{rw}'].value}")
                    self.ui.agen_label_3.show()
                    self.ui.client_name_out.setText(f" Клиент: {sheet[f'I{rw}'].value}")
                    self.ui.client_name_out.show()
                    self.ui.client_name_val_out.setText(f" Сумма заказа клиента: {sheet[f'J{rw}'].value}")
                    self.ui.client_name_val_out.show()

                if status == '1':
                    self.all_hide()

                    self.ui.key_inp_label.show()
                    self.ui.key_inp.show()
                    self.ui.key_inp_button.show()
                    self.ui.key_val_label.show()
                    key_status = 'Активен'
                    key_status_button = 'Деактивировать'
                    self.ui.key_status_label.setText(f'{key_status}')
                    self.ui.key_status_label.show()
                    self.ui.status_button.setText(key_status_button)
                    self.ui.status_button.show()
                    self.ui.agen_label.setText(f' Агент: {sheet[f"F{rw}"].value}')
                    self.ui.agen_label.show()
                    self.ui.date_label.setText(f" Дата генерации: {sheet[f'C{rw}'].value}")
                    self.ui.date_label.show()
                    self.ui.deadline_label.setText(f' Срок действия: {sheet[f"E{rw}"].value}')
                    self.ui.deadline_label.show()
                    self.ui.status_button.clicked.connect(lambda: get_client_name(rw, key_status))
                    # [change_status(rw, 0), self.main_func()]
                elif status == '0':
                    self.all_hide()

                    self.ui.key_inp_label.show()
                    self.ui.key_inp.show()
                    self.ui.key_inp_button.show()
                    self.ui.key_val_label.show()
                    key_status = 'Неактивен'
                    key_status_button = 'Активировать'
                    self.ui.key_status_label.setText(f'{key_status}')
                    self.ui.key_status_label.show()
                    self.ui.status_button.setText(key_status_button)
                    self.ui.status_button.show()
                    self.ui.status_button.clicked.connect(lambda: [change_status(rw, 1), self.key_checker(), name_of_agent(key_status, rw)])

    def update_label(self, value):
        self.ui.deadline_label_2.setText(str(value))

    def update_xls(self, rows):
        need_date = datetime.today() + relativedelta(months=int(self.ui.deadline_slider.value()))
        file = 'C:\Program Files (x86)\Key Master by sshunin\data.xlsx'
        wookbook = openpyxl.load_workbook(file)
        sheet = wookbook.active
        sheet[f'F{rows}'] = str(self.ui.agent_name_inp.text())
        sheet[f'E{rows}'] = f'{datetime.today().strftime("%d.%m.%Y")} - {need_date.strftime("%d.%m.%Y")}'
        sheet[f'H{rows}'] = int(self.ui.agent_name_inp_value.text())
        sheet[f'D{rows}'] = 1
        wookbook.save(file)
        self.log_saver(f'[{datetime.today().strftime("%H:%M:%S")}] Key {self.ui.key_inp.text()} was successfully activated; Agent name: {str(self.ui.agent_name_inp.text())}, Validity period: {datetime.today().strftime("%d.%m.%Y")} - {need_date.strftime("%d.%m.%Y")}\n',
            False)


app = QtWidgets.QApplication([])
application = Generator()
application.show()

sys.exit(app.exec_())